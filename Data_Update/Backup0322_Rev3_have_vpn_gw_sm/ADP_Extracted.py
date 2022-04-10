# -*- coding: utf-8 -*-

#eric-pc-firewall-coordinator-at 这个很特殊，resources还要下一级

import yaml
import openpyxl
from string import ascii_uppercase
from openpyxl.styles import Font
import os


class Extract_Data(): #类，为了取值。各个def可以换顺序

    def __init__(self,path,subfolders,filename,excel_path,product): # 起始，传参数进来，self代表着新建的实例对象
        self.path = path
        self.subfolders = subfolders
        self.filename = filename
        self.excel_path = excel_path
        self.product=product

    # def openexcel(self): #打开Excel
    #     file = openpyxl.Workbook() # 调用Workbook这个类

    def extract(self,subfolder,absolutepath): #打开并读取yaml file，并调用process函数。data里存的是yaml file全文，比如Log Transformer全文
        with open(absolutepath) as file:
            data = yaml.safe_load(file) # 读取yaml file，读取出来是嵌套的字典
            self.process(subfolder,data)



    def loopfile(self): # 遍历subfolders，得到每个subfolder的绝对路径，并调用extract函数
        for subfolder in self.subfolders:
            if subfolder != 'eric-pc-sm':
                absolutepath = self.path+"/"+subfolder+"/"+self.filename
                self.extract(subfolder,absolutepath)
                # print(subfolder)
            elif subfolder == 'eric-pc-sm':
                subfolder_charts = os.listdir(self.path+"/"+ 'eric-pc-sm/charts')
                print (subfolder_charts)
                for subsubfolder in subfolder_charts:
                    absolutepath = self.path+"/"+subfolder+"/" + 'charts/' + subsubfolder +'/' +self.filename
                    self.extract(subsubfolder, absolutepath)
                    print (subsubfolder)
            # elif subfolder == 'eric-pc-mm':


    def action(self):
        # self.workbook.close() 怎么强制关闭Excel呢？
        self.workbook = openpyxl.Workbook()
        # self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet(title='ADP',index=0)
        # self.worksheet.sheet_properties.tabColor='0000FF' #sheet tab color
        self.worksheet.append(["Chart Name","Container Name","Requests_CPU","Requests_Memory","Requests_Storage","Limits_CPU","Limits_Memory","Limits_Storage","Replicas"])
        self.loopfile()
        print("Done for getting ADP data.")
        self.workbook.remove(self.workbook['Sheet'])
        # # self.worksheet.column_dimensions['A'].auto_size = True #如何自适应Excel宽度？
        # # self.worksheet.column_dimensions['B'].width =200
        for column in ascii_uppercase:  # 改Excel宽度
            if column == 'A':
                self.worksheet.column_dimensions[column].width = 36
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                self.worksheet.column_dimensions[column].width = 30
            else:
                self.worksheet.column_dimensions[column].width = 12
        # self.worksheet.cell(1, 1).font = Font(name='Arial', size=10) #改这一个可以成功
        for i in range(1,self.worksheet.max_row + 1): #改Excel字体
            for j in range(1,self.worksheet.max_column+1):
                self.worksheet.cell(i,j).font = Font(name='Arial', size=10)
        self.workbook.save(filename=self.excel_path) #按照excel_path，保存路径。这里的filename=可要可不要


    # def return_value(self,dictionary,first_level,second_level):
    #     try:
    #         return dictionary[first_level][second_level] # return的什么？比如resource["requests"]["cpu"]，return的是250ms, 在这里，resource是{'requests': {'cpu': '250m', 'memory': '2Gi', 'ephemeral-storage': None}, 'limits': {'cpu': '1000m', 'memory': '2Gi', 'ephemeral-storage': None}}
    #     except:
    #         return None


    def get_replicas(self,data,key_list_replica_value, key_list_replicaCount_value):
        # 用instance检测a是否为字典
        if isinstance(data, dict): #data是一个subfolder，即一个pod的字典数据
            # print(type(data))
            Final_value=[]
            for i in range(len(data)):
                # print(i)
                temp_key = list(data.keys())[i]  # 将所有的key转化为了一个list，再是list列表的第几个元素
                temp_value = data[temp_key]
                # print(temp_key)
                # print(list(data.keys()))
                # print(temp_value)
                if temp_key == 'replicas':
                    key_list_replica_value.append(temp_value)
                    # print('key_list_replica_value is: %s' % key_list_replica_value)
                if temp_key == 'replicaCount':
                    key_list_replicaCount_value.append(temp_value)
                    # print('key_list_replicaCount_value is: %s' % key_list_replicaCount_value)
                # break

                self.get_replicas(temp_value, key_list_replica_value, key_list_replicaCount_value)

                if len(key_list_replica_value) == 0 and len(key_list_replicaCount_value) == 0:
                    Final_value = ['No replica info']

                elif len(key_list_replica_value) == 1 and len(key_list_replicaCount_value) == 0:
                    Final_value = key_list_replica_value

                elif len(key_list_replica_value) == 0 and len(key_list_replicaCount_value) == 1:
                    Final_value = key_list_replicaCount_value

                elif len(key_list_replica_value) == 1 and len(key_list_replicaCount_value) == 1:
                    Final_value = ['Two replica info. One is replicas and one is replicaCount. Check']

                elif len(key_list_replica_value) == 2:
                    Final_value = ['Two replicas. Check']

                elif len(key_list_replica_value) > 2:
                    Final_value = ['More than two replicas. Check']

                elif len(key_list_replicaCount_value) == 2:
                    Final_value = ['Two replicaCounts. Check']

                elif len(key_list_replicaCount_value) > 2:
                    Final_value = ['More than two replicaCounts. Check']

                # print("Final_value is %s" % Final_value)
                # ret = self.get_replicas(temp_value,key_list_replica_value, key_list_replicaCount_value)
            # print('Final End Final Value %s' % Final_value)
            return Final_value #return相当于循环结束

        else:
            pass



    # def get_replicas(self,data, key_list_replica_value, key_list_replicaCount_value, default=None):
    #     Final_value =[]
    #     for key, value in data.items():
    #         if key == 'replicas':  # 找到一次就退出了
    #             key_list_replica_value.append(value)
    #             Final_value = value
    #         elif key == 'replicaCount':
    #             key_list_replicaCount_value.append(value)
    #             Final_value = value
    #         else:
    #             if isinstance(value, dict):  # 是字典就执行
    #                 ret = self.get_replicas(value, key_list_replica_value, key_list_replicaCount_value, default)
    #                 if len(key_list_replica_value) == 0 and len(key_list_replicaCount_value) == 0:
    #                     Final_value = ['No replica info']
    #                 elif len(key_list_replica_value) == 1 and len(key_list_replicaCount_value) == 0:
    #                     Final_value = key_list_replica_value
    #                 elif len(key_list_replica_value) == 0 and len(key_list_replicaCount_value) == 1:
    #                     Final_value = key_list_replicaCount_value
    #                 elif len(key_list_replica_value) == 1 and len(key_list_replicaCount_value) == 1:
    #                     Final_value = ['Two replica info. One is replicas and one is replicaCount. Check']
    #                 elif len(key_list_replica_value) == 2:
    #                     Final_value = ['Two replicas. Check']
    #                 elif len(key_list_replica_value) > 2:
    #                     Final_value = ['More than two replicas. Check']
    #                 elif len(key_list_replicaCount_value) == 2:
    #                     Final_value = ['Two replicaCounts. Check']
    #                 elif len(key_list_replicaCount_value) > 2:
    #                     Final_value = ['More than two replicaCounts. Check']
    #                 if ret is not default:  # 如果ret与default = None不等，表示找到replica或replicaCount，ret作为返回值往上返回。
    #                     return ret
    #                     return default
    #                     return Final_value
    #             else:
    #                 pass







    def return_value(self,dictionary,levels): #传进来是limits/requests的字典，并指定上下级的list
        thedict = dictionary
        thelevel = levels[::-1] # 先把各个level翻转，比如["requests","cpu"]变成["cpu","requests"]
        while thelevel: #thelevel字典的值不为空时
            try:
                thedict = thedict[thelevel.pop()] #一层一层弹出时，先弹出requests,对应的值也就是{'cpu': '250m', 'memory': '2Gi', 'ephemeral-storage': None}
            except:
                thedict = None
        return thedict


    def get_resources(self,subfolder, data,replica_data):
        if self.product == 'PCG' and subfolder == 'eric-pc-firewall-coordinator-at':
            at_level_data = data['at']
            resources = at_level_data['resources']
        elif self.product == 'PCC' and subfolder == 'eric-pc-vpn-gw':
            eric_pc_vpn_gw_level_data = data['eric-pc-vpn-gw']
            resources = eric_pc_vpn_gw_level_data['resources']
        else:
            if "resources" in data:
                resources = data["resources"] #取得键值为resources的value的值
                # print(resources)
            else:
                data=[]
                data.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data.append("")  # 第二个resource不存在，存空
                data.append("")
                data.append("")
                data.append("")
                data.append("")
                data.append("")
                data.append("")
                data.append(replica_data[0]) #最后一列是replica info
                # print(data)
                self.worksheet.append(data)
                resources=[] #这句话必须得写，如果是空的话，得有空列表，下面的for循环才能用

        #再得到resources的数据
        for container in resources:
            # print(container)
            global data_list
            data_list = []
            data_list.append(subfolder) # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
            data_list.append(container) # 第二列是resource，就是resources里的每个resource，比如logtransformer
            data_list.append(self.return_value(resources[container],["requests","cpu"])) #调用return_value函数，输入为一个字典和不定的众多levels
            data_list.append(self.return_value(resources[container], ["requests", "memory"]))
            data_list.append(self.return_value(resources[container], ["requests", "ephemeral-storage"]))
            data_list.append(self.return_value(resources[container], ["limits", "cpu"]))
            data_list.append(self.return_value(resources[container], ["limits", "memory"]))
            data_list.append(self.return_value(resources[container], ["limits", "ephemeral-storage"]))
            data_list.append(replica_data[0])
            print(data_list)
            self.worksheet.append(data_list)

        # return data_list


    def process(self, subfolder, data):  # 从yaml file里取出resource和replica的值
        if subfolder != 'eric-data-search-engine': # 不是search engine
            # 先得到replica的值
            # print(subfolder)
            global key_list_replica_value
            key_list_replica_value = []
            global key_list_replicaCount_value
            key_list_replicaCount_value = []
            # global  Final_value
            # Final_value = []
            # print(subfolder)
            #先得到replica info
            replica_data = self.get_replicas(data=data, key_list_replica_value=key_list_replica_value,key_list_replicaCount_value=key_list_replicaCount_value)
            #再把replica info传参，再去得到resources
            self.get_resources(subfolder,data, replica_data)
        else:
            ingest_replica, master_replica, data_replica = self.get_replicas_search_engine(data)
            # print(ingest_replica)
            # print(master_replica)
            # print(data_replica)
            self.get_resources_search_engine(subfolder,data,ingest_replica,master_replica,data_replica)




    def get_replicas_search_engine(self,data):
        if isinstance(data, dict):
            for i in range(len(data)):
                temp_key = list(data.keys())[i]  # 将key转化为了一个list，再是list列表的第几个元素
                temp_value = data[temp_key]
                if temp_key == 'replicaCount':
                    replica_search_engine = temp_value #replica_search_engine此时是字典{'ingest': 1, 'master': 3, 'data': 2}
                    ingest_replica = replica_search_engine['ingest']
                    master_replica = replica_search_engine['master']
                    data_replica = replica_search_engine['data']
            return ingest_replica,master_replica,data_replica
        else:
            pass


    def get_resources_search_engine(self,subfolder,data,ingest_replica,master_replica,data_replica):
        resources = data["resources"]
        for resource in resources:
            if resource == 'ingest':
                data_list = []
                data_list.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append(resource)  # 第二列是resource，就是resources里的每个resource，比如logtransformer
                data_list.append(self.return_value(resources[resource], ["requests", "cpu"]))  # 调用return_value函数，输入为一个字典和不定的众多levels
                data_list.append(self.return_value(resources[resource], ["requests", "memory"]))
                data_list.append(self.return_value(resources[resource], ["requests", "ephemeral-storage"]))
                data_list.append(self.return_value(resources[resource], ["limits", "cpu"]))
                data_list.append(self.return_value(resources[resource], ["limits", "memory"]))
                data_list.append(self.return_value(resources[resource], ["limits", "ephemeral-storage"]))
                data_list.append(ingest_replica)
                self.worksheet.append(data_list)
            elif resource == 'master':
                data_list = []
                data_list.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append(resource)  # 第二列是resource，就是resources里的每个resource，比如logtransformer
                data_list.append(self.return_value(resources[resource], ["requests", "cpu"]))  # 调用return_value函数，输入为一个字典和不定的众多levels
                data_list.append(self.return_value(resources[resource], ["requests", "memory"]))
                data_list.append(self.return_value(resources[resource], ["requests", "ephemeral-storage"]))
                data_list.append(self.return_value(resources[resource], ["limits", "cpu"]))
                data_list.append(self.return_value(resources[resource], ["limits", "memory"]))
                data_list.append(self.return_value(resources[resource], ["limits", "ephemeral-storage"]))
                data_list.append(master_replica)
                self.worksheet.append(data_list)
            elif resource == 'data':
                data_list = []
                data_list.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append(resource)  # 第二列是resource，就是resources里的每个resource，比如logtransformer
                data_list.append(self.return_value(resources[resource], ["requests", "cpu"]))  # 调用return_value函数，输入为一个字典和不定的众多levels
                data_list.append(self.return_value(resources[resource], ["requests", "memory"]))
                data_list.append(self.return_value(resources[resource], ["requests", "ephemeral-storage"]))
                data_list.append(self.return_value(resources[resource], ["limits", "cpu"]))
                data_list.append(self.return_value(resources[resource], ["limits", "memory"]))
                data_list.append(self.return_value(resources[resource], ["limits", "ephemeral-storage"]))
                data_list.append(data_replica)
                self.worksheet.append(data_list)
            else:
                data_list = []
                data_list.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append(resource)  # 第二列是resource，就是resources里的每个resource，比如logtransformer
                data_list.append(self.return_value(resources[resource], ["requests", "cpu"]))  # 调用return_value函数，输入为一个字典和不定的众多levels
                data_list.append(self.return_value(resources[resource], ["requests", "memory"]))
                data_list.append(self.return_value(resources[resource], ["requests", "ephemeral-storage"]))
                data_list.append(self.return_value(resources[resource], ["limits", "cpu"]))
                data_list.append(self.return_value(resources[resource], ["limits", "memory"]))
                data_list.append(self.return_value(resources[resource], ["limits", "ephemeral-storage"]))
                data_list.append('No replica info')
                self.worksheet.append(data_list)


