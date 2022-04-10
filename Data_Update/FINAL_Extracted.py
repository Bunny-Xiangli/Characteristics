# -*- coding: utf-8 -*-

#eric-pc-firewall-coordinator-at 这个很特殊，resources还要下一级

import yaml
import openpyxl
import ADP_Extracted
import PCG_Extracted
import USER_Extracted
import FINAL_Extracted
from string import ascii_uppercase
from openpyxl.styles import colors, fills, Font, PatternFill


class Final_Extract_Data(): #类，为了取值，各个def可以换顺序

    def __init__(self,excel_path): # 起始，传参数进来，self代表着新建的实例对象
        self.excel_path = excel_path

    def action(self,excel_path):
        self.workbook = openpyxl.load_workbook(excel_path)
        self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet('FINAL')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        self.worksheet = self.workbook.create_sheet('MISMATCH')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        # self.getfile()
        # self.replace(PCG_Extracted.Extract_Data_2nd_PCG.process(),USER_Extracted.Extract_Data_3rd_USER.process()) #想得到PCG和User里的Excel结果
        # self.replace(ADP_Extracted.Extract_Data.process(),PCG_Extracted.Extract_Data_2nd_PCG.process())
        worksheet_ADP = self.workbook['ADP']
        worksheet_PCX = self.workbook['PCX']
        worksheet_FINAL = self.workbook['FINAL']
        worksheet_CHECK = self.workbook['MISMATCH']

        # 把ADP的全部cell copy到Final sheet里
        for row in worksheet_ADP:
            for cell in row:
                worksheet_FINAL[cell.coordinate].value=cell.value

        if 'USER' in self.workbook.sheetnames: #for PCG
            worksheet_USER = self.workbook['USER']
            self.replace(low_priority=worksheet_FINAL, high_priority=worksheet_PCX, worksheet_CHECK=worksheet_CHECK)
            self.replace_user(low_priority=worksheet_FINAL, high_priority=worksheet_USER,worksheet_CHECK=worksheet_CHECK)
        else: #for PCC
            self.replace(low_priority=worksheet_FINAL, high_priority=worksheet_PCX,worksheet_CHECK=worksheet_CHECK)
        # print(worksheet_CHECK.max_row)

        #改单位和数值，与CPI F4 Match
        self.remove_units_and_adjust_format()

        # 找到需要check的empty container
        worksheet_FINAL.cell(row=1, column=10).value = 'Check Emtpy Data'
        # worksheet_FINAL.cell(row=1, column=10).font = Font(name='Arial', size=10, color='C00000')
        for i in range(1, worksheet_FINAL.max_row + 1):
            if worksheet_FINAL.cell(row=i, column=2).value is None:
                worksheet_FINAL.cell(row=i, column=10).value = "Container and Resource"  # container是空，则该数据需要检查
                worksheet_FINAL.cell(row=i, column=10).font = Font(name='Arial', size=10, color='C00000')
            # elif worksheet_FINAL.cell(row=i, column=3).value is None and \
            #     worksheet_FINAL.cell(row=i, column=4).value is None\
            #     worksheet_FINAL.cell(row=i, column=5).value is None\
            #     worksheet_FINAL.cell(row=i, column=6).value is None\
            #     worksheet_FINAL.cell(row=i, column=7).value is None\
            #     worksheet_FINAL.cell(row=i, column=8).value is None\
            #     worksheet_FINAL.cell(row=i, column=9).value is None:
            #     worksheet_FINAL.cell(row=i, column=10).value = "Check the empty empty CPU, memory, and storage"
            #     worksheet_FINAL.cell(row=i, column=10).font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='C00000')
            elif worksheet_FINAL.cell(row=i, column=9).value is None:
                worksheet_FINAL.cell(row=i, column=10).value = "Replica"  # replica是空，则该数据需要检查
                worksheet_FINAL.cell(row=i, column=10).font = Font(name='Arial', size=10, color='C00000')
            else:
                pass


        # 改Excel宽度
        for sheet in ['FINAL', 'MISMATCH']:
            ws = self.workbook[sheet]
            for column in ascii_uppercase:
                if column == 'A':
                    ws.column_dimensions[column].width = 30
                    # self.worksheet.cell.font = Font(name='Arial',size=10)
                elif column == 'B':
                    ws.column_dimensions[column].width = 25
                elif column == 'I' or column == 'Q':
                    ws.column_dimensions[column].width = 10
                else:
                    ws.column_dimensions[column].width = 16

        # for column in ascii_uppercase:
        #     if column == 'A':
        #         worksheet_CHECK.column_dimensions[column].width = 30
        #         # self.worksheet.cell.font = Font(name='Arial',size=10)
        #     elif column == 'B':
        #         worksheet_CHECK.column_dimensions[column].width = 25
        #     elif column == 'I' or column == 'Q':
        #         worksheet_CHECK.column_dimensions[column].width = 10
        #     else:
        #         worksheet_CHECK.column_dimensions[column].width = 16




        # 改Excel字体
        for i in range(1, worksheet_FINAL.max_row +1 ):
            for j in range(1, worksheet_FINAL.max_column+1):
                worksheet_FINAL.cell(i, j).font = Font(name='Arial', size=10)
        # 改Excel字体
        for i in range(1, worksheet_CHECK.max_row +1):
            for j in range(1, worksheet_CHECK.max_column +1):
                worksheet_CHECK.cell(i, j).font = Font(name='Arial', size=10)




        # #自动filter worksheet_FINAL这一列
        # worksheet_FINAL.auto_filter.ref = worksheet_FINAL.dimensions

        print ('Done for merging data. You can close GUI.')
        # print (worksheet_CHECK.max_row)
        self.workbook.save(filename=self.excel_path)  # 按照excel_path，保存路径。这里的filename=可要可不要
        # openpyxl.load_workbook(excel_path) #怎么自动打开Excel




    def replace(self,low_priority,high_priority,worksheet_CHECK=None):
        if high_priority.cell(2,1).value is None: # 如果该表没有数据，就pass
            pass
        else:
            data_list_low =[]
            data_list_high =[]
            #得到low_priority的podname list
            for i in range(1,low_priority.max_row+1):
                low_priority.cell(row=i, column=15).value = str(low_priority.cell(row=i, column=1).value) + '___' + str(low_priority.cell(row=i, column=2).value)
                data_list_low.append(low_priority.cell(row=i, column=15).value)
            # print(data_list_low)

            #得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value = str(high_priority.cell(row=i, column=1).value) + '___' + str(high_priority.cell(row=i, column=2).value)
                data_list_high.append(high_priority.cell(row=i, column=15).value)
            # print(data_list_high)

            #比较然后replace
            for i in range(1, high_priority.max_row + 1):
                each_pod_high = high_priority.cell(row=i, column=1).value
                to_be_checked = []
                #万一high_pod_container不在low list里，则应该单独看，加到CHECK worksheet
                if high_priority.cell(row=i, column=15).value not in data_list_low:
                    # low_priority.cell(row=i,column=10).value = 'Check for non-exist pod/container' #不能，需要单独创建一章表，因为老的表里不会有这些，是在high_priority里的
                    to_be_checked.append(each_pod_high)
                    to_be_checked.append(high_priority.cell(row=i, column=2).value)
                    to_be_checked.append(high_priority.cell(row=i, column=3).value)
                    to_be_checked.append(high_priority.cell(row=i, column=4).value)
                    to_be_checked.append(high_priority.cell(row=i, column=5).value)
                    to_be_checked.append(high_priority.cell(row=i, column=6).value)
                    to_be_checked.append(high_priority.cell(row=i, column=7).value)
                    to_be_checked.append(high_priority.cell(row=i, column=8).value)
                    to_be_checked.append(high_priority.cell(row=i, column=9).value)
                    # print(to_be_checked)
                    worksheet_CHECK.append(to_be_checked)
                else:
                    for j in range(1, low_priority.max_row + 1):
                        each_pod_low = low_priority.cell(row=j, column=1).value
                        if each_pod_high == 'Chart Name' or each_pod_low == 'Chart Name':
                            pass
                        elif high_priority.cell(row=i, column=15).value == low_priority.cell(row=j,column=15).value:
                            if high_priority.cell(row=i, column=3).value is not None:
                                low_priority.cell(row=j, column=3).value = high_priority.cell(row=i,column=3).value
                                low_priority.cell(row=j, column=3).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色 #B8CCE4 浅灰蓝色
                            else:
                                pass
                            if high_priority.cell(row=i, column=4).value is not None:
                                low_priority.cell(row=j, column=4).value = high_priority.cell(row=i, column=4).value
                                low_priority.cell(row=j, column=4).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=5).value is not None:
                                low_priority.cell(row=j, column=5).value = high_priority.cell(row=i, column=5).value
                                low_priority.cell(row=j, column=5).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=6).value is not None:
                                low_priority.cell(row=j, column=6).value = high_priority.cell(row=i, column=6).value
                                low_priority.cell(row=j, column=6).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=7).value is not None:
                                low_priority.cell(row=j, column=7).value = high_priority.cell(row=i, column=7).value
                                low_priority.cell(row=j, column=7).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=8).value is not None:
                                low_priority.cell(row=j, column=8).value = high_priority.cell(row=i, column=8).value
                                low_priority.cell(row=j, column=8).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=9).value is not None:
                                low_priority.cell(row=j, column=9).value = high_priority.cell(row=i, column=9).value
                                low_priority.cell(row=j, column=9).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                        else:
                            pass

            #看不懂了，啥意思？第15列是pod name_container name，最后把这一列赋空，相当于删掉了
            for i in range(1, low_priority.max_row + 1):
                low_priority.cell(row=i, column=15).value =''

            # 得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value =''


    def replace_user(self,low_priority,high_priority,worksheet_CHECK=None): #只有颜色不同，赋予了User黄色
        if high_priority.cell(2,1).value is None:
            pass
        else:
            data_list_low =[]
            data_list_high =[]
            #得到low_priority的podname list
            for i in range(1,low_priority.max_row+1):
                low_priority.cell(row=i, column=15).value = str(low_priority.cell(row=i, column=1).value) + '___' + str(low_priority.cell(row=i, column=2).value)
                data_list_low.append(low_priority.cell(row=i, column=15).value)
            # print(data_list_low)

            #得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value = str(high_priority.cell(row=i, column=1).value) + '___' + str(high_priority.cell(row=i, column=2).value)
                data_list_high.append(high_priority.cell(row=i, column=15).value)
            # print(data_list_high)

            #比较然后replace
            for i in range(1, high_priority.max_row + 1):
                each_pod_high = high_priority.cell(row=i, column=1).value
                to_be_checked = []
                flag=0
                #万一high_pod_container不在low list里，则应该单独看，加到CHECK worksheet
                if high_priority.cell(row=i, column=15).value not in data_list_low:
                    flag=flag+1
                    to_be_checked.append(each_pod_high)
                    to_be_checked.append(high_priority.cell(row=i, column=2).value)
                    to_be_checked.append(high_priority.cell(row=i, column=3).value)
                    to_be_checked.append(high_priority.cell(row=i, column=4).value)
                    to_be_checked.append(high_priority.cell(row=i, column=5).value)
                    to_be_checked.append(high_priority.cell(row=i, column=6).value)
                    to_be_checked.append(high_priority.cell(row=i, column=7).value)
                    to_be_checked.append(high_priority.cell(row=i, column=8).value)
                    to_be_checked.append(high_priority.cell(row=i, column=9).value)
                    # print(to_be_checked)
                    worksheet_CHECK.append(to_be_checked)
                else:
                    for j in range(1, low_priority.max_row + 1):
                        each_pod_low = low_priority.cell(row=j, column=1).value
                        if each_pod_high == 'Chart Name' or each_pod_low == 'Chart Name':
                            pass
                        elif high_priority.cell(row=i, column=15).value == low_priority.cell(row=j,column=15).value:
                            if high_priority.cell(row=i, column=3).value is not None:
                                low_priority.cell(row=j, column=3).value = high_priority.cell(row=i,column=3).value
                                low_priority.cell(row=j, column=3).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色 C4D79B 浅灰绿色
                            else:
                                pass
                            if high_priority.cell(row=i, column=4).value is not None:
                                low_priority.cell(row=j, column=4).value = high_priority.cell(row=i, column=4).value
                                low_priority.cell(row=j, column=4).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=5).value is not None:
                                low_priority.cell(row=j, column=5).value = high_priority.cell(row=i, column=5).value
                                low_priority.cell(row=j, column=5).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=6).value is not None:
                                low_priority.cell(row=j, column=6).value = high_priority.cell(row=i, column=6).value
                                low_priority.cell(row=j, column=6).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=7).value is not None:
                                low_priority.cell(row=j, column=7).value = high_priority.cell(row=i, column=7).value
                                low_priority.cell(row=j, column=7).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=8).value is not None:
                                low_priority.cell(row=j, column=8).value = high_priority.cell(row=i, column=8).value
                                low_priority.cell(row=j, column=8).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=9).value is not None:
                                low_priority.cell(row=j, column=9).value = high_priority.cell(row=i, column=9).value
                                low_priority.cell(row=j, column=9).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                        else:
                            pass

            for i in range(1, low_priority.max_row + 1):
                low_priority.cell(row=i, column=15).value =''

            # 得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value =''



    def remove_units_and_adjust_format(self):
        for sheet in ['FINAL','MISMATCH']:
            ws = self.workbook[sheet]
            # print(ws)

            # 想在最前面加一列Pod_Container，用作对比CPI
            ws.insert_cols(idx=11, amount=1)
            ws.cell(1, 11).value = 'For_CPI_Pod_Container'
            for i in range (1,ws.max_row+1):
                if i==1:
                    ws.cell(1, 11).value = 'For_CPI_Pod_Container'
                elif ws.cell(i, 1).value is not None or ws.cell(i, 2).value is not None:
                    ws.cell(i, 11).value = str(ws.cell(i, 1).value) + "_" + str(ws.cell(i, 2).value)
                else:
                    pass

            # 第一列Requests_CPU_Format，按后面带啥单位执行
            ws.insert_cols(idx=12, amount=1)  # 在replica后增加一列
            for i in range(1, ws.max_row + 1):
                Requests_CPU_data = ws.cell(i, 3).value
                if Requests_CPU_data == 'Requests_CPU':
                    ws.cell(i, 12).value = Requests_CPU_data + '_CPI'  # 处理新增列的抬头，免得标题重复
                elif str(Requests_CPU_data).endswith('m'):
                    str_data = str(Requests_CPU_data).replace('m', '')  # 去掉单位m
                    float_data = float(str_data)
                    float_data = float_data / 1000
                    ws.cell(i, 12).value = float_data
                elif Requests_CPU_data is not None:
                    ws.cell(i, 12).value = float(Requests_CPU_data)
                else:
                    pass
                    # ws.cell(i, 11).value ='Wrong(RC)'
                    # ws.cell(i,11).fill = PatternFill(fill_type='solid', fgColor="FB5D66")
                # print (i)
                # print (ws.cell(i,11).value)
                # str.replace(r'm','').astype(float)


            #第二列Requests_Memory__Format，按后面带啥单位执行
            ws.insert_cols(idx=13, amount=1)
            for i in range(1,ws.max_row +1):
                Requests_Memory_data = ws.cell(i, 4).value
                if Requests_Memory_data == 'Requests_Memory':
                    ws.cell(i,13).value = Requests_Memory_data + '_CPI'  #处理新增列的抬头，免得标题重复
                elif str(Requests_Memory_data).endswith('Mi') :
                    str_data_Mi = str(Requests_Memory_data).replace('Mi', '')  # 去掉单位Mi
                    ws.cell(i, 13).value = float(str_data_Mi) #
                elif str(Requests_Memory_data).endswith('M'):
                    str_data_M = str(Requests_Memory_data).replace('M', '')  # 去掉单位Mi
                    ws.cell(i, 13).value = float(str_data_M)  #
                elif str(Requests_Memory_data).endswith('Gi'):
                    str_data_Gi = str(Requests_Memory_data).replace('Gi', '')  # 去掉单位Gi
                    ws.cell(i, 13).value = float (str_data_Gi) * 1000
                elif Requests_Memory_data is None:
                    pass
                else:
                    pass
                    # ws.cell(i, 12).value ='Wrong(RM)'
                    # ws.cell(i, 12).fill = PatternFill(fill_type='solid', fgColor="FB5D66")


            # 第三列Requests_Storage_Format
            ws.insert_cols(idx=14, amount=1)
            for i in range(1, ws.max_row + 1):
                Requests_Storage_data = ws.cell(i, 5).value
                if Requests_Storage_data == 'Requests_Storage':
                    ws.cell(i, 14).value = Requests_Storage_data + '_CPI'  # 处理新增列的抬头，免得标题重复
                elif str(Requests_Storage_data).endswith('Mi'):
                    str_data_Mi = str(Requests_Storage_data).replace('Mi', '')  # 去掉单位Mi
                    ws.cell(i, 14).value = float(str_data_Mi)
                elif str(Requests_Storage_data).endswith('Gi'):
                    str_data_Gi = str(Requests_Storage_data).replace('Gi', '')  # 去掉单位Gi
                    ws.cell(i, 14).value = float(str_data_Gi) * 1000
                elif Requests_Storage_data is None:
                    pass
                else:
                    pass
                    # ws.cell(i, 13).value ='Wrong(RS)'
                    # ws.cell(i, 13).fill = PatternFill(fill_type='solid', fgColor="FB5D66")



            #第四列Limits_CPU_Format，按后面带啥单位执行
            ws.insert_cols(idx=15, amount=1) #在replica后增加一列
            for i in range(1,ws.max_row +1):
                Limits_CPU_data = ws.cell(i, 6).value
                if Limits_CPU_data == 'Limits_CPU':
                    ws.cell(i,15).value = Limits_CPU_data + '_CPI'  #处理新增列的抬头，免得标题重复
                elif str(Limits_CPU_data).endswith('m'):
                    str_data = str(Limits_CPU_data).replace('m', '')  # 去掉单位m
                    float_data = float(str_data)
                    float_data = float_data / 1000
                    ws.cell(i, 15).value = float_data
                elif Limits_CPU_data is not None:
                    ws.cell(i, 15).value = float(Limits_CPU_data)
                else:
                    pass
                    # ws.cell(i, 14).value ='Wrong(LC)'
                    # ws.cell(i, 14).fill = PatternFill(fill_type='solid', fgColor="FB5D66")


            #第五列Limits_Memory_Format，按后面带啥单位执行
            ws.insert_cols(idx=16, amount=1)
            for i in range(1, ws.max_row + 1):
                Limits_Memory_data = ws.cell(i, 7).value
                if Limits_Memory_data == 'Limits_Memory':
                    ws.cell(i, 16).value = Limits_Memory_data + '_CPI'  # 处理新增列的抬头，免得标题重复
                elif str(Limits_Memory_data).endswith('Mi'):
                    str_data_Mi = str(Limits_Memory_data).replace('Mi', '')  # 去掉单位Mi
                    ws.cell(i, 16).value = int(str_data_Mi)
                elif str(Limits_Memory_data).endswith('M'):
                    str_data_M = str(Limits_Memory_data).replace('M', '')  # 去掉单位Mi
                    ws.cell(i, 16).value = int(str_data_M)
                elif str(Limits_Memory_data).endswith('Gi'):
                    str_data_Gi = str(Limits_Memory_data).replace('Gi', '')  # 去掉单位Gi
                    ws.cell(i, 16).value = float(str_data_Gi) * 1000
                elif Limits_Memory_data is None:
                    pass
                else:
                    pass
                    # ws.cell(i, 15).value = 'Wrong(LM)'
                    # ws.cell(i, 15).fill = PatternFill(fill_type='solid', fgColor="FB5D66")


            # 第六列Limits_Storage_Format
            ws.insert_cols(idx=17, amount=1)
            for i in range(1, ws.max_row + 1):
                Limits_Storage_data = ws.cell(i, 8).value
                if Limits_Storage_data == 'Limits_Storage':
                    ws.cell(i, 17).value = Limits_Storage_data + '_CPI'  # 处理新增列的抬头，免得标题重复
                elif str(Limits_Storage_data).endswith('Mi'):
                    str_data_Mi = str(Limits_Storage_data).replace('Mi', '')  # 去掉单位Mi
                    ws.cell(i, 17).value = int(str_data_Mi)
                elif str(Limits_Storage_data).endswith('Gi'):
                    str_data_Gi = str(Limits_Storage_data).replace('Gi', '')  # 去掉单位Gi
                    ws.cell(i, 17).value = float(str_data_Gi) * 1000
                elif Limits_Storage_data is None:
                    pass
                else:
                    pass
                    # ws.cell(i, 16).value = 'Wrong(LS)'
                    # ws.cell(i,16).fill = PatternFill(fill_type='solid', fgColor="FB5D66")

            ws.insert_cols(idx=18, amount=1)
            for i in range(1,ws.max_row+1):
                Replicas = ws.cell(i,9).value
                if Replicas == "Replicas":
                    ws.cell(i,18).value = Replicas + '_CPI'
                else:
                    ws.cell(i,18).value=Replicas

