# -*- coding: utf-8 -*-

#读取yaml file
import yaml
import openpyxl
import json
import os
import openpyxl
from openpyxl import Workbook, load_workbook

#for yaml
# f=open(r"C:\Users\lxi\OneDrive - Sigma Technology\Desktop\eric-pc-gateway-1.47.0-44\eric-pc-gateway-1.47.0-44\eric-pc-gateway\supporting-files\example-config\values_test.yaml")
# data=yaml.safe_load_all(f)
# # print(data)
#
# for d in data:
#     print (d)
#
# # for key in data:
# #     print (key)






# def __init__(self, path):  # 起始，传参数进来，self代表着新建的实例对象
#     self.path = path
#
#
# def write_chart_name_to_excel(self,path):
#     workbook = self.openpyxl.Workbook()
#     worksheet = self.workbook.active
#     worksheet.append(["Name", "2nd half"])
#     run(self,path)
#     workbook.save(filename=r'C:\XL\Sigma\Trainings\Python\Others\For Liangrui_handle_json\output.xlsx')  # 按照excel_path，保存路径。这里的filename=可要可不要
#
# def run(self,path):
#     f = open(path)
#     raw_data = json.load(f)
#     data_2nd = raw_data.get('data')
#     ericsson_2nd = data_2nd.get('ericsson-pm:pm')
#     group_2nd = ericsson_2nd.get('group')
#     # print (group_2nd)
#     for item in group_2nd:
#         list1 = []
#         # name= group_2nd.get('name')
#         # print(item)
#         name=item.get('name')
#         measurement_type=item.get('measurement-type')
#         # print (item.get('measurement-type'))
#         # print (measurement_type['name'])
#
#         # list1.append(name)
#         list1.append(measurement_type)
#         print (list1)
#         self.worksheet.append(list1)
#         # list.append()
#
#
# if __name__ == '__main__':
#     write_chart_name_to_excel(r"C:\XL\Sigma\Trainings\Python\Others\For Liangrui_handle_json\bulk_report_1.json")



workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.append(["Name", "2nd half"])

f = open(r"C:\XL\Sigma\Trainings\Python\Others\For Liangrui_handle_json\bulk_report_2.json")
raw_data = json.load(f)
data_2nd = raw_data.get('data')
ericsson_2nd = data_2nd.get('ericsson-pm:pm')
group_2nd = ericsson_2nd.get('group')
# print (group_2nd)
count=0

for item in group_2nd:
    list1 = []
    count = count + 1
    # name= group_2nd.get('name')
    # print(item)
    name=item.get('name')
    measurement_type=item.get('measurement-type')
    # print (item.get('measurement-type'))
    # print (measurement_type['name'])

    # list1.append(name)

    list1.append(name)
    print (list1)

    worksheet.append(list1)




workbook.save(filename=r'C:\XL\Sigma\Trainings\Python\Others\For Liangrui_handle_json\output.xlsx')
