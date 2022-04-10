# -*- coding: utf-8 -*-

#eric-pc-firewall-coordinator-at 这个很特殊，resources还要下一级

import yaml
import openpyxl
from string import ascii_uppercase
from openpyxl.styles import Font


class Extract_Data(): #类，为了取值。各个def可以换顺序

    def __init__(self,path,subfolders,filename,excel_path): # 起始，传参数进来，self代表着新建的实例对象
        self.path = path
        self.subfolders = subfolders
        self.filename = filename
        self.excel_path = excel_path

    # def openexcel(self): #打开Excel
    #     file = openpyxl.Workbook() # 调用Workbook这个类

    def extract(self,subfolder,absolutepath): #打开并读取yaml file，并调用process函数。data里存的是yaml file全文，比如Log Transformer全文
        with open(absolutepath) as file:
            data = yaml.safe_load(file) # 读取yaml file，读取出来是嵌套的字典
            self.process(subfolder,data)
            # print(data)

    def loopfile(self): # 遍历subfolders，得到每个subfolder的绝对路径，并调用extract函数
        for subfolder in self.subfolders:
            absolutepath = self.path+"/"+subfolder+"/"+self.filename
            self.extract(subfolder,absolutepath)

    def action(self):
        # self.workbook.close() 怎么强制关闭Excel呢？
        self.workbook = openpyxl.Workbook()
        # self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet(title='ADP',index=0)
        # self.worksheet.sheet_properties.tabColor='0000FF' #sheet tab color
        self.worksheet.append(["Chart Name","Container Name","Requests_CPU","Requests_Memory","Requests_Storage","Limits_CPU","Limits_Memory","Limits_Storage","Replicas"])
        self.loopfile()
        print("Done for getting ADP data.")
        self.workbook.remove(self.workbook['Sheet'])
        # # self.worksheet.column_dimensions['A'].auto_size = True #如何自适应Excel宽度？
        # # self.worksheet.column_dimensions['B'].width =200
        for column in ascii_uppercase:  # 改Excel宽度
            if column == 'A':
                self.worksheet.column_dimensions[column].width = 36
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                self.worksheet.column_dimensions[column].width = 30
            else:
                self.worksheet.column_dimensions[column].width = 12
        # self.worksheet.cell(1, 1).font = Font(name='Arial', size=10) #改这一个可以成功
        for i in range(1,self.worksheet.max_row + 1): #改Excel字体
            for j in range(1,self.worksheet.max_column+1):
                self.worksheet.cell(i,j).font = Font(name='Arial', size=10)
        self.workbook.save(filename=self.excel_path) #按照excel_path，保存路径。这里的filename=可要可不要


    # def return_value(self,dictionary,first_level,second_level):
    #     try:
    #         return dictionary[first_level][second_level] # return的什么？比如resource["requests"]["cpu"]，return的是250ms, 在这里，resource是{'requests': {'cpu': '250m', 'memory': '2Gi', 'ephemeral-storage': None}, 'limits': {'cpu': '1000m', 'memory': '2Gi', 'ephemeral-storage': None}}
    #     except:
    #         return None

    def return_value(self,dictionary,levels):
        thedict = dictionary
        thelevel = levels[::-1] # 先把各个level翻转，比如["requests","cpu"]变成["cpu","requests"]
        while thelevel:
            try:
                thedict = thedict[thelevel.pop()] #一层一层弹出时，先弹出requests,对应的值也就是{'cpu': '250m', 'memory': '2Gi', 'ephemeral-storage': None}
            except:
                thedict = None
        return thedict


#eric-pc-firewall-coordinator-at，这个resources藏得比较深，没有读出来
    def process(self,subfolder,data): #从yaml file里取出resource和replica的值
        # should meet the data structure
        replica = data.get("replicaCount",None)
        replica_2nd = data.get ("replicas",None)
        if "resources" in data: #写这段儿是让resources不要返回None?
            resources = data["resources"]
            # print(resources)
            for resource in resources:
                data_list = []
                data_list.append(subfolder) # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append(resource) # 第二列是resource，就是resources里的每个resource，比如logtransformer
                data_list.append(self.return_value(resources[resource],["requests","cpu"])) #调用return_value函数，输入为一个字典和不定的众多levels
                data_list.append(self.return_value(resources[resource], ["requests", "memory"]))
                data_list.append(self.return_value(resources[resource], ["requests", "ephemeral-storage"]))
                data_list.append(self.return_value(resources[resource], ["limits", "cpu"]))
                data_list.append(self.return_value(resources[resource], ["limits", "memory"]))
                data_list.append(self.return_value(resources[resource], ["limits", "ephemeral-storage"]))
                # data_list.append(self.return_value(resources[resource],"requests","cpu"))
                # data_list.append(self.return_value(resources[resource], "requests", "memory"))
                # data_list.append(self.return_value(resources[resource], "requests", "ephemeral-storage"))
                # data_list.append(self.return_value(resources[resource], "limits", "cpu"))
                # data_list.append(self.return_value(resources[resource], "limits", "memory"))
                # data_list.append(self.return_value(resources[resource], "limits", "ephemeral-storage"))
                # data_list.append(resources[resource].get("requests",None).get("cpu",None))
                # data_list.append(resources[resource].get("requests",None).get("memory",None))
                # data_list.append(resources[resource].get("requests",None).get("ephemeral-storage",None))
                # data_list.append(resources[resource].get("limits",None).get("cpu",None))
                # data_list.append(resources[resource].get("limits",None).get("memory",None))
                # data_list.append(resources[resource].get("limits",None).get("ephemeral-storage",None))
                # data_list.append(replica)
                # data_list.append(replica_2nd)
                # data_list.append(replica+replica_2nd) #有None，不能加
                if replica is None:
                    data_list.append(replica_2nd)
                else:
                    data_list.append(replica)
                print(data_list)
                self.worksheet.append(data_list)
        else:
            if replica != None or replica_2nd != None:
                data_list = []
                data_list.append(subfolder)  # 第一列是subfolder，就是subfolders列表里的每一个值，也就是pod name，比如eric-log-transformer
                data_list.append("")  # 第二个resource不存在，存空
                data_list.append("")
                data_list.append("")
                data_list.append("")
                data_list.append("")
                data_list.append("")
                data_list.append("")
                # data_list.append(replica)
                # data_list.append(replica_2nd)
                if replica is None:
                    data_list.append(replica_2nd)
                else:
                    data_list.append(replica)
                print(data_list)
                self.worksheet.append(data_list)
            else:
                return








