# -*- coding: utf-8 -*-

#eric-pc-firewall-coordinator-at 这个很特殊，resources还要下一级

import yaml
import openpyxl
import ADP_Extracted
import PCG_Extracted
import USER_Extracted
import FINAL_Extracted
from string import ascii_uppercase
from openpyxl.styles import colors, fills, Font, PatternFill


class Final_Extract_Data(): #类，为了取值，各个def可以换顺序

    def __init__(self,excel_path): # 起始，传参数进来，self代表着新建的实例对象
        self.excel_path = excel_path

    def action(self,excel_path):
        self.workbook = openpyxl.load_workbook(excel_path)
        self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet('FINAL')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        self.worksheet = self.workbook.create_sheet('MISMATCH')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        # self.getfile()
        # self.replace(PCG_Extracted.Extract_Data_2nd_PCG.process(),USER_Extracted.Extract_Data_3rd_USER.process()) #想得到PCG和User里的Excel结果
        # self.replace(ADP_Extracted.Extract_Data.process(),PCG_Extracted.Extract_Data_2nd_PCG.process())
        worksheet_ADP = self.workbook['ADP']
        worksheet_PCX = self.workbook['PCX']
        worksheet_FINAL = self.workbook['FINAL']
        worksheet_CHECK = self.workbook['MISMATCH']


        # 把ADP的全部cell copy到Final sheet里
        for row in worksheet_ADP:
            for cell in row:
                worksheet_FINAL[cell.coordinate].value=cell.value

        if 'USER' in self.workbook.sheetnames: #for PCG
            worksheet_USER = self.workbook['USER']
            self.replace(low_priority=worksheet_FINAL, high_priority=worksheet_PCX, worksheet_CHECK=worksheet_CHECK)
            self.replace_user(low_priority=worksheet_FINAL, high_priority=worksheet_USER,worksheet_CHECK=worksheet_CHECK)
        else: #for PCC
            self.replace(low_priority=worksheet_FINAL, high_priority=worksheet_PCX,worksheet_CHECK=worksheet_CHECK)



        # 改Excel宽度
        for column in ascii_uppercase:
            if column == 'A':
                worksheet_FINAL.column_dimensions[column].width = 30
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                worksheet_FINAL.column_dimensions[column].width = 25
            elif column == 'I':
                worksheet_FINAL.column_dimensions[column].width = 10
            else:
                worksheet_FINAL.column_dimensions[column].width = 12

        for column in ascii_uppercase:
            if column == 'A':
                worksheet_CHECK.column_dimensions[column].width = 30
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                worksheet_CHECK.column_dimensions[column].width = 25
            elif column == 'I':
                worksheet_CHECK.column_dimensions[column].width = 10
            else:
                worksheet_CHECK.column_dimensions[column].width = 12

        # 改Excel字体
        for i in range(1, worksheet_FINAL.max_row +1 ):
            for j in range(1, worksheet_FINAL.max_column+1):
                worksheet_FINAL.cell(i, j).font = Font(name='Arial', size=10)
        # 改Excel字体
        for i in range(1, worksheet_CHECK.max_row +1):
            for j in range(1, worksheet_CHECK.max_column +1):
                worksheet_CHECK.cell(i, j).font = Font(name='Arial', size=10)


        # 找到需要check的empty container
        # worksheet_FINAL.cell(row=1, column=10).value = 'To Check'
        # worksheet_FINAL.cell(row=1, column=10).font = Font(name='Arial', size=10, color='C00000')
        for i in range(1, worksheet_FINAL.max_row + 1):
            if worksheet_FINAL.cell(row=i, column=2).value is None:
                worksheet_FINAL.cell(row=i, column=10).value = "Container and Resource"
                worksheet_FINAL.cell(row=i, column=10).font = Font(name='Arial', size=10,color='C00000')
            # elif worksheet_FINAL.cell(row=i, column=3).value is None and \
            #     worksheet_FINAL.cell(row=i, column=4).value is None\
            #     worksheet_FINAL.cell(row=i, column=5).value is None\
            #     worksheet_FINAL.cell(row=i, column=6).value is None\
            #     worksheet_FINAL.cell(row=i, column=7).value is None\
            #     worksheet_FINAL.cell(row=i, column=8).value is None\
            #     worksheet_FINAL.cell(row=i, column=9).value is None:
            #     worksheet_FINAL.cell(row=i, column=10).value = "Check the empty empty CPU, memory, and storage"
            #     worksheet_FINAL.cell(row=i, column=10).font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color='C00000')
            elif worksheet_FINAL.cell(row=i, column=9).value is None:
                worksheet_FINAL.cell(row=i, column=10).value = "Replica"
                worksheet_FINAL.cell(row=i, column=10).font = Font(name='Arial', size=10,color='C00000')
            else:
                pass

        # #自动filter worksheet_FINAL这一列
        # worksheet_FINAL.auto_filter.ref = worksheet_FINAL.dimensions

        print ('Done for merging data. You can close GUI.')
        self.workbook.save(filename=self.excel_path)  # 按照excel_path，保存路径。这里的filename=可要可不要
        # openpyxl.load_workbook(excel_path) #怎么自动打开Excel






    def replace(self,low_priority,high_priority,worksheet_CHECK=None):
        if high_priority.cell(2,1).value is None:
            pass
        else:
            data_list_low =[]
            data_list_high =[]
            #得到low_priority的podname list
            for i in range(1,low_priority.max_row+1):
                low_priority.cell(row=i, column=15).value = str(low_priority.cell(row=i, column=1).value) + '___' + str(low_priority.cell(row=i, column=2).value)
                data_list_low.append(low_priority.cell(row=i, column=15).value)
            # print(data_list_low)

            #得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value = str(high_priority.cell(row=i, column=1).value) + '___' + str(high_priority.cell(row=i, column=2).value)
                data_list_high.append(high_priority.cell(row=i, column=15).value)
            # print(data_list_high)

            #比较然后replace
            for i in range(1, high_priority.max_row + 1):
                each_pod_high = high_priority.cell(row=i, column=1).value
                to_be_checked = []
                #万一high_pod_container不在low list里，则应该单独看，加到CHECK worksheet
                if high_priority.cell(row=i, column=15).value not in data_list_low:
                    # low_priority.cell(row=i,column=10).value = 'Check for non-exist pod/container' #不能，需要单独创建一章表，因为老的表里不会有这些，是在high_priority里的
                    to_be_checked.append(each_pod_high)
                    to_be_checked.append(high_priority.cell(row=i, column=2).value)
                    to_be_checked.append(high_priority.cell(row=i, column=3).value)
                    to_be_checked.append(high_priority.cell(row=i, column=4).value)
                    to_be_checked.append(high_priority.cell(row=i, column=5).value)
                    to_be_checked.append(high_priority.cell(row=i, column=6).value)
                    to_be_checked.append(high_priority.cell(row=i, column=7).value)
                    to_be_checked.append(high_priority.cell(row=i, column=8).value)
                    to_be_checked.append(high_priority.cell(row=i, column=9).value)
                    # print(to_be_checked)
                    worksheet_CHECK.append(to_be_checked)
                else:
                    for j in range(1, low_priority.max_row + 1):
                        each_pod_low = low_priority.cell(row=j, column=1).value
                        if each_pod_high == 'Chart Name' or each_pod_low == 'Chart Name':
                            pass
                        elif high_priority.cell(row=i, column=15).value == low_priority.cell(row=j,column=15).value:
                            if high_priority.cell(row=i, column=3).value is not None:
                                low_priority.cell(row=j, column=3).value = high_priority.cell(row=i,column=3).value
                                low_priority.cell(row=j, column=3).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色 #B8CCE4 浅灰蓝色
                            else:
                                pass
                            if high_priority.cell(row=i, column=4).value is not None:
                                low_priority.cell(row=j, column=4).value = high_priority.cell(row=i, column=4).value
                                low_priority.cell(row=j, column=4).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=5).value is not None:
                                low_priority.cell(row=j, column=5).value = high_priority.cell(row=i, column=5).value
                                low_priority.cell(row=j, column=5).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=6).value is not None:
                                low_priority.cell(row=j, column=6).value = high_priority.cell(row=i, column=6).value
                                low_priority.cell(row=j, column=6).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=7).value is not None:
                                low_priority.cell(row=j, column=7).value = high_priority.cell(row=i, column=7).value
                                low_priority.cell(row=j, column=7).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=8).value is not None:
                                low_priority.cell(row=j, column=8).value = high_priority.cell(row=i, column=8).value
                                low_priority.cell(row=j, column=8).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=9).value is not None:
                                low_priority.cell(row=j, column=9).value = high_priority.cell(row=i, column=9).value
                                low_priority.cell(row=j, column=9).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                            else:
                                pass
                        else:
                            pass

            for i in range(1, low_priority.max_row + 1):
                low_priority.cell(row=i, column=15).value =''

            # 得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value =''


    def replace_user(self,low_priority,high_priority,worksheet_CHECK=None): #只有颜色不同，赋予了User黄色
        if high_priority.cell(2,1).value is None:
            pass
        else:
            data_list_low =[]
            data_list_high =[]
            #得到low_priority的podname list
            for i in range(1,low_priority.max_row+1):
                low_priority.cell(row=i, column=15).value = str(low_priority.cell(row=i, column=1).value) + '___' + str(low_priority.cell(row=i, column=2).value)
                data_list_low.append(low_priority.cell(row=i, column=15).value)
            # print(data_list_low)

            #得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value = str(high_priority.cell(row=i, column=1).value) + '___' + str(high_priority.cell(row=i, column=2).value)
                data_list_high.append(high_priority.cell(row=i, column=15).value)
            # print(data_list_high)

            #比较然后replace
            for i in range(1, high_priority.max_row + 1):
                each_pod_high = high_priority.cell(row=i, column=1).value
                to_be_checked = []
                flag=0
                #万一high_pod_container不在low list里，则应该单独看，加到CHECK worksheet
                if high_priority.cell(row=i, column=15).value not in data_list_low:
                    flag=flag+1
                    to_be_checked.append(each_pod_high)
                    to_be_checked.append(high_priority.cell(row=i, column=2).value)
                    to_be_checked.append(high_priority.cell(row=i, column=3).value)
                    to_be_checked.append(high_priority.cell(row=i, column=4).value)
                    to_be_checked.append(high_priority.cell(row=i, column=5).value)
                    to_be_checked.append(high_priority.cell(row=i, column=6).value)
                    to_be_checked.append(high_priority.cell(row=i, column=7).value)
                    to_be_checked.append(high_priority.cell(row=i, column=8).value)
                    to_be_checked.append(high_priority.cell(row=i, column=9).value)
                    # print(to_be_checked)
                    worksheet_CHECK.append(to_be_checked)
                else:
                    for j in range(1, low_priority.max_row + 1):
                        each_pod_low = low_priority.cell(row=j, column=1).value
                        if each_pod_high == 'Chart Name' or each_pod_low == 'Chart Name':
                            pass
                        elif high_priority.cell(row=i, column=15).value == low_priority.cell(row=j,column=15).value:
                            if high_priority.cell(row=i, column=3).value is not None:
                                low_priority.cell(row=j, column=3).value = high_priority.cell(row=i,column=3).value
                                low_priority.cell(row=j, column=3).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色 C4D79B 浅灰绿色
                            else:
                                pass
                            if high_priority.cell(row=i, column=4).value is not None:
                                low_priority.cell(row=j, column=4).value = high_priority.cell(row=i, column=4).value
                                low_priority.cell(row=j, column=4).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=5).value is not None:
                                low_priority.cell(row=j, column=5).value = high_priority.cell(row=i, column=5).value
                                low_priority.cell(row=j, column=5).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=6).value is not None:
                                low_priority.cell(row=j, column=6).value = high_priority.cell(row=i, column=6).value
                                low_priority.cell(row=j, column=6).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=7).value is not None:
                                low_priority.cell(row=j, column=7).value = high_priority.cell(row=i, column=7).value
                                low_priority.cell(row=j, column=7).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=8).value is not None:
                                low_priority.cell(row=j, column=8).value = high_priority.cell(row=i, column=8).value
                                low_priority.cell(row=j, column=8).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                            if high_priority.cell(row=i, column=9).value is not None:
                                low_priority.cell(row=j, column=9).value = high_priority.cell(row=i, column=9).value
                                low_priority.cell(row=j, column=9).fill = PatternFill(fill_type='solid',fgColor="FFFF00")  # 如果覆盖了就变色
                            else:
                                pass
                        else:
                            pass

            for i in range(1, low_priority.max_row + 1):
                low_priority.cell(row=i, column=15).value =''

            # 得到high_priority的podname list
            for i in range(1, high_priority.max_row + 1):
                high_priority.cell(row=i, column=15).value =''
