# -*- coding: utf-8 -*-

import os
import openpyxl
from openpyxl import Workbook, load_workbook

class Get_chart_name_and_write_to_excel():

    def __init__(self,path): # 起始，传参数进来，self代表着新建的实例对象
        self.path = path

    def get_chart_name(self):
        # 为什么不能写成 def get_chart_name(self,path)，不是要把path传进来么？报错：TypeError: get_chart_name() missing 1 required positional argument: 'path'
        chart_names = os.listdir(self.path)
        # print(chart_names)
        return chart_names

    def write_chart_name_to_excel(self):

        # chart_names = os.listdir(path)
        chart_names = self.get_chart_name()

        wb = Workbook()
        ws = wb.active
        ws.title='Changed Data'

        # ws.append(chart_names)

        for i in range(0, len(chart_names)): # 注意：range的时候，是len (chart_names)-1，而达不到len(chart_names)
            each_chart_name = chart_names[i]
            ws.cell(row=i+1, column=2).value = chart_names[i]
            # ws.append(each_chart_name)
            print(each_chart_name)
            print (i)


        # for chart_name in chart_names: # 这里要怎么写才对呢？不能写成chart_names，例如：  for chart_name in chart_names:
        #     ws.append(chart_name)

        wb.save(filename=r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\Get_chart_name_and_write_to_excel.xlsx')

        # read_excel = openpyxl.load_workbook(r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\Get_chart_name_and_write_to_excel.xlsx')
        # print (read_excel)

if __name__ == '__main__':
    path = r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\eric-pc-gateway-1.47.0-44\eric-pc-gateway-1.47.0-44\eric-pc-gateway\charts'
    # path = r'C:\XL\Sigma\Trainings\Python\Fapiao'
    # Get_chart_name_and_write_to_excel.get_chart_name(path)
    Get_chart_name_and_write_to_excel(path).write_chart_name_to_excel()


#
# #模仿的这里的代码：https://blog.csdn.net/GYK0812/article/details/104380629，也弄不出来
#
# chart_names = ['eric-cm-mediator', 'eric-cm-yang-provider', 'eric-cnom-document-database-mg', 'eric-cnom-server',
#                'eric-ctrl-bro', 'eric-data-coordinator-zk', 'eric-data-distributed-coordinator-ed',
#                'eric-data-document-database-pg', 'eric-data-message-bus-kf', 'eric-data-object-storage-mn',
#                'eric-data-search-engine', 'eric-data-search-engine-curator', 'eric-data-sftp-server',
#                'eric-fh-alarm-handler', 'eric-fh-snmp-alarm-provider', 'eric-lm-combined-server', 'eric-log-shipper',
#                'eric-log-transformer', 'eric-odca-diagnostic-data-collector', 'eric-pc-firewall-coordinator-at',
#                'eric-pc-kvdb-rd-operator', 'eric-pc-kvdb-rd-server', 'eric-pc-networking-cm-agent',
#                'eric-pc-routing-aggregator', 'eric-pc-routing-engine', 'eric-pc-routing-engine-api',
#                'eric-pc-up-br-agent', 'eric-pc-up-data-plane', 'eric-pc-up-partitioner', 'eric-pc-up-pfcp-endpoint',
#                'eric-pc-up-service-chaining-cm-agent', 'eric-pc-up-timer-wheel', 'eric-pc-up-user-plane-cm-agent',
#                'eric-pm-bulk-reporter', 'eric-pm-resource-monitor', 'eric-pm-server', 'eric-probe-event-report-broker',
#                'eric-sec-admin-user-management', 'eric-sec-certm', 'eric-sec-key-management', 'eric-sec-ldap-server',
#                'eric-sec-sip-tls', 'eric-sw-inventory-manager']
#
# # chart_names = ["eric-cm-mediator", "eric-cm-yang-provider"]
#
# print(chart_names)
#
# wb = Workbook()
# ws = wb.active
# ws.title = 'Get_chart_name_and_write_to_excel'
#
# for chart_name in chart_names:
#     ws.append(chart_name)
# # # wb.save(r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop')




