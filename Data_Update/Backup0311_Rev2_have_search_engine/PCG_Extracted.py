# -*- coding: utf-8 -*-

import yaml
import openpyxl
import pandas as pd
from string import ascii_uppercase
from openpyxl.styles import Font

class Extract_Data_2nd_PCG():

    def __init__(self, path, filename,excel_path):  # 起始，传参数进来，self代表着新建的实例对象
        self.path = path
        self.filename = filename
        self.excel_path = excel_path

    def extract(self, absolutepath):  # 打开并读取yaml file，并调用process函数。data里存的是yaml file全文，比如Log Transformer全文
        with open(absolutepath) as file:
            data = yaml.safe_load(file)  # 读取yaml file，读取出来是嵌套的字典
            self.maybepodname(data) #传入PCG values.yaml的字典

    def maybepodname(self, data):
        podnamelist_PCG = []
        for everypodname in data:  # 得到的就是data里面的key值，也就是podname
            podnamelist_PCG.append(everypodname)  # podnamelist_PCG，就是最后带有所有的podname的list
            # print(everypodname)
        # podnamelist_PCG.remove("tags")
        # podnamelist_PCG.remove("global")
        # podnamelist_PCG.remove("m2m")
        # podnamelist_PCG.remove("productInfoStatus")
        # podnamelist_PCG.remove("labels")
        # print(podnamelist_PCG) #podnamelist_PCG是个列表
        self.process(podnamelist_PCG, data)

    def action(self,excel_path):
        # self.workbook = openpyxl.Workbook()
        # self.worksheet = self.workbook.active
        self.workbook = openpyxl.load_workbook(excel_path)
        self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet('PCX')
        # self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        self.getfile()
        print("Done for getting PCG data.")
        for column in ascii_uppercase:  # 改Excel宽度
            if column == 'A':
                self.worksheet.column_dimensions[column].width = 36
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                self.worksheet.column_dimensions[column].width = 30
            else:
                self.worksheet.column_dimensions[column].width = 12
        for i in range(1, 200):  # 改Excel字体
            for j in range(1, 15):
                self.worksheet.cell(i, j).font = Font(name='Arial', size=10)
        self.workbook.save(filename=self.excel_path)  # 按照excel_path，保存路径。这里的filename=可要可不要



        # read_excel = pd.read_excel(r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\eric-pc-gateway-1.47.0-44\Output.xlsx')
        # data = pd.DataFrame(columns=["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage",
        #              "Limits_CPU", "Limits_Memory", "Limits_Storage", "Replicas"])
        # self.getfile()
        # data.to_excel(r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\eric-pc-gateway-1.47.0-44\Output.xlsx','PCG')




    def getfile(self):  # 得到values.yaml的绝对路径
        absolutepath = self.path + "/" + self.filename
        self.extract(absolutepath)

    def return_value(self, dictionary, levels):
        thedict = dictionary
        thelevel = levels[::-1]  # 先把各个level翻转，比如["requests","cpu"]变成["cpu","requests"]
        while thelevel:
            try:
                thedict = thedict[
                    thelevel.pop()]  # 一层一层弹出时，先弹出requests,对应的值也就是{'cpu': '250m', 'memory': '2Gi', 'ephemeral-storage': None}
            except:
                thedict = None #为什么要用except?
        return thedict #只返回了一个值，比如cpu对应的值

        # def find_keywords(self,data_len,podnamelist_PCG,data): #尝试把带有resources的关键词搜索出来
        #     queue = [data]
        #     while len(data) > 0:
        #         left_data = queue.pop()
        #         for key,value in left_data.items():

    def process(self, podnamelist_PCG, data):  # 从yaml file里取出resource和replica的值
        for everypodname in podnamelist_PCG:
            if everypodname != "eric-data-search-engine": #serch_engine的replica有两层级
                everypodname_2ndhalf = data.get(everypodname)  # get前半截的key，可以得到后半截数据, 比如得到eric-log-transformer: 后半截数据
                # replica_1 = everypodname_2ndhalf.get('replicaCount') #没有把replica写在这里抓出来是因为可能抓出来的是一个string，比如enabled，就会报错，因为string没有get方法
                # replica_2 = everypodname_2ndhalf.get('replicas')
                #本来准备写下面一段，结果发现有一样的，获得replica info，核对下没有问题，故注释掉
                # for key in everypodname_2ndhalf:
                #     if key == 'replicas':
                #         replica = everypodname_2ndhalf['replicas']
                #         print ('replicas')
                #         print (everypodname)
                #         print(replica)
                #     elif key == 'replicaCount':
                #         replica = everypodname_2ndhalf['replicaCount']
                #         print('replicaCount')
                #         print(everypodname)
                #         print(replica)
                #     else:
                #         pass


                if "resources" in everypodname_2ndhalf:  # 这里查出来的结果不全呢？答案：是因为之前最后用了return，直接跳出了循环
                    resources_details = everypodname_2ndhalf["resources"] #也是取字典key后半截数据，和get有什么区别呢？区别在于，用get即使没有找到也不会报错
                    replica_1 = everypodname_2ndhalf.get('replicaCount',None) #没有把replica写在这里抓出来是因为可能抓出来的是一个srting，比如enabled，就会报错，因为string没有get方法
                    replica_2 = everypodname_2ndhalf.get('replicas',None)
                    if ("requests" or "limits") in resources_details: # eric-cnom-document-database-mg，这个就很奇怪，没有container这一层，在pod下直接跟着resources，然后requests
                        self.get_detailed_resources(replica_1, replica_2, resources_details, everypodname)  # 传的是requests/limits这一层
                        # print (everypodname)
                    else:
                        for container in resources_details: #到了container级别，比如resources的下一级，logtransformer
                            # resources_details[container] # key对应value
                            self.get_detailed_resources(replica_1, replica_2, resources_details[container], everypodname, container)  # 传的是requests/limits这一层，从requests/limits开始
                elif "replicaCount" in everypodname_2ndhalf: #eric-sec-ldap-server只有replicaCount，没有resources
                    replica_value = everypodname_2ndhalf.get('replicaCount', None)
                    data_list = []
                    data_list.append(everypodname)
                    data_list.append("")  # 第二个container，存空
                    data_list.append("")  # 第三个Requests，存空
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append(replica_value)
                    data_list.append("")
                    # print(data_list)
                    self.worksheet.append(data_list)
                elif "replicas" in everypodname_2ndhalf:
                    replica_value_2nd = everypodname_2ndhalf.get('replicas', None)
                    data_list = []
                    data_list.append(everypodname)
                    data_list.append("")  # 第二个container，存空
                    data_list.append("")  # 第三个Requests，存空
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append("")
                    data_list.append(replica_value_2nd)
                    # print(data_list)
                    self.mworksheet.append(data_list)
                else:
                    continue  # 中断这一次的for循环，进入下一次的for循环。我第一次用成了return，则是跳出了这个for循环
            elif everypodname == "eric-data-search-engine":
                everypodname_search_engine_2ndhalf = data.get(everypodname)
                if "replicaCount" in everypodname_search_engine_2ndhalf:
                    replica_level = everypodname_search_engine_2ndhalf['replicaCount']
                    replica_ingest = replica_level.get('ingest',None)
                    # print(replica_level)
                    # print(replica_ingest)
                    replica_master = replica_level.get('master', None)
                    replica_data = replica_level.get('data', None)
                    for replica in replica_level:
                        if replica != 'ingest' or replica != 'master' or replica != 'data':
                            replica = 'No replica in ingest, master, or data. Check'

                if "resources" in everypodname_search_engine_2ndhalf:
                    resources_details = everypodname_search_engine_2ndhalf.get("resources", None)
                    for container in resources_details:
                        data_list_search_engine = []
                        container_level_2nd_half = resources_details.get(container, None)
                        # print(container_level_2nd_half)
                        data_list_search_engine.append(everypodname)
                        data_list_search_engine.append(container)
                        # for req_or_lim in container_level_2nd_half: #不能要for 循环，要不跑了两遍
                        # 取Requests CPU的值：
                        dict_Req = container_level_2nd_half.get('requests', None)
                        if dict_Req is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Req_cpu = dict_Req.get('cpu', None)
                            if dict_Req_cpu is not None:
                                data_list_search_engine.append(dict_Req_cpu)
                            else:
                                data_list_search_engine.append('')

                        # 取Requests Memory值：
                        if dict_Req is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Req_mem = dict_Req.get('memory', None)
                            if dict_Req_mem is not None:
                                data_list_search_engine.append(dict_Req_mem)
                            else:
                                data_list_search_engine.append('')

                        # 取Requests Memory Storage值：
                        if dict_Req is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Req_stor = dict_Req.get('ephemeral-storage', None)
                            if dict_Req_stor is not None:
                                data_list_search_engine.append(dict_Req_stor)
                            else:
                                data_list_search_engine.append('')

                        # 取Limits CPU的值：
                        dict_Lim = container_level_2nd_half.get('limits',
                                                                None)  # 是limits对应的数据，比如{'memory': '3Gi', 'cpu': '750m'}
                        if dict_Lim is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Lim_cpu = dict_Lim.get('cpu', None)
                            if dict_Lim_cpu is not None:
                                data_list_search_engine.append(dict_Lim_cpu)
                            else:
                                data_list_search_engine.append('')

                        # #取Limits Memory的值：
                        if dict_Lim is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Lim_mem = dict_Lim.get('memory', None)
                            if dict_Lim_mem is not None:
                                data_list_search_engine.append(dict_Lim_mem)
                            else:
                                data_list_search_engine.append('')

                        # #取Limits Stor的值：
                        if dict_Lim is None:
                            data_list_search_engine.append('')
                        else:
                            dict_Lim_stor = dict_Lim.get('ephemeral-storage', None)
                            if dict_Lim_stor is not None:
                                data_list_search_engine.append(dict_Lim_stor)
                            else:
                                data_list_search_engine.append('')

                        # 追加一列replica信息
                        if container == 'ingest':
                            data_list_search_engine.append(replica_ingest)
                        elif container == 'master':
                            data_list_search_engine.append(replica_master)
                        elif container == 'data':
                            data_list_search_engine.append(replica_data)
                        else:
                            data_list_search_engine.append(replica)

                        self.worksheet.append(data_list_search_engine)
                        print(data_list_search_engine)

            else:
                continue

    def get_detailed_resources(self, replica_1, replica_2,resources_level1, everypodname, container=None):  # 万一没有传参数，container=None, 默认的container值是None
        data_resource_list = []
        data_resource_list.append(everypodname)  # 第一列是pod name，比如eric-log-transformer
        data_resource_list.append(container)  # 第二列是resource，就是resources里的每个resource，比如logtransformer
        data_resource_list.append(self.return_value(resources_level1, ["requests", "cpu"]))  # 调用return_value函数，输入为一个字典和不定的众多levels
        data_resource_list.append(self.return_value(resources_level1, ["requests", "memory"]))
        data_resource_list.append(self.return_value(resources_level1, ["requests", "ephemeral-storage"]))
        data_resource_list.append(self.return_value(resources_level1, ["limits", "cpu"]))
        data_resource_list.append(self.return_value(resources_level1, ["limits", "memory"]))
        data_resource_list.append(self.return_value(resources_level1, ["limits", "ephemeral-storage"]))
        # data_resource_list.append(replica_1)
        # data_resource_list.append(replica_2)
        if replica_1 is None:
            data_resource_list.append(replica_2)
        else:
            data_resource_list.append(replica_1)
        # return data_list
        # print(data_resource_list)
        self.worksheet.append(data_resource_list)








