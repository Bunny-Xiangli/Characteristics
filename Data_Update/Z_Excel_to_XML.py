import openpyxl


workbook = openpyxl.load_workbook(r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\NACM\nacm_rule_name_update_in_PCC_2nd_Renamed.xlsx')

worksheet = workbook.active

filename = r'C:\Users\lxi\OneDrive - Sigma Technology\Desktop\NACM\PCC_output_Renamed.txt'

with open (filename,'w') as f: #如果filename不在则自动创建，w表示写数据
    for i in range(1,worksheet.max_row+1):
        f.write('<row>\n')
        for j in range (1,worksheet.max_column+1):
            data = []
            data.append(worksheet.cell(i,j).value)
            # data.append('<row>\n')
            # data.append('<entry>\n')
            # data.append('<p>')
            # data.append(worksheet.cell(i,j).value)
            # data.append('</p>\n')
            # data.append('</entry>\n')
            # data.append(('/row>' + '\n'))
            # print (data)

            f.write('  <entry>\n')
            f.write('    <p>')
            if data[0] is not None:
                f.write(data[0])
            else:
                f.write('')
            f.write('</p>\n')
            f.write('  </entry>\n')
        f.write('</row>\n')

f.close()





