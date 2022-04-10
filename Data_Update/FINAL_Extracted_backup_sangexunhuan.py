# -*- coding: utf-8 -*-

#eric-pc-firewall-coordinator-at 这个很特殊，resources还要下一级

import yaml
import openpyxl
import ADP_Extracted
import ADP_settings
import PCG_Extracted
import PCG_settings
import USER_Extracted
import USER_settings
import FINAL_Extracted
import FINAL_settings
from openpyxl.styles import PatternFill
from string import ascii_uppercase
from openpyxl.styles import Font


class Final_Extract_Data(): #类，为了取值，各个def可以换顺序

    def __init__(self,path, excel_path): # 起始，传参数进来，self代表着新建的实例对象
        self.path = path
        self.excel_path = excel_path

    def action(self,excel_path):
        self.workbook = openpyxl.load_workbook(excel_path)
        self.worksheet = self.workbook.active
        self.worksheet = self.workbook.create_sheet('FINAL')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        self.worksheet = self.workbook.create_sheet('CHECK')
        self.worksheet.sheet_properties.tabColor = '1072BA'
        self.worksheet.append(
            ["Chart Name", "Container Name", "Requests_CPU", "Requests_Memory", "Requests_Storage", "Limits_CPU",
             "Limits_Memory", "Limits_Storage", "Replicas"])
        # self.getfile()
        # self.replace(PCG_Extracted.Extract_Data_2nd_PCG.process(),USER_Extracted.Extract_Data_3rd_USER.process()) #想得到PCG和User里的Excel结果
        # self.replace(ADP_Extracted.Extract_Data.process(),PCG_Extracted.Extract_Data_2nd_PCG.process())
        worksheet_ADP = self.workbook['ADP']
        worksheet_PCX = self.workbook['PCX']
        worksheet_USER = self.workbook['USER']
        worksheet_FINAL = self.workbook['FINAL']
        worksheet_CHECK = self.workbook['CHECK']

        # 把ADP的全部cell copy到Final sheet里
        # for row in worksheet_ADP:
        #     for cell in row:
        #         worksheet_FINAL[cell.coordinate].value=cell.value

        # 把PCG的全部cell copy到Final sheet里
        for row in worksheet_PCX:
            for cell in row:
                worksheet_FINAL[cell.coordinate].value=cell.value

        # self.replace (low_priority=worksheet_FINAL,high_priority=worksheet_PCX)
        self.replace(low_priority=worksheet_FINAL, high_priority=worksheet_USER)

        # 改Excel宽度
        for column in ascii_uppercase:
            if column == 'A':
                worksheet_FINAL.column_dimensions[column].width = 36
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                worksheet_FINAL.column_dimensions[column].width = 30
            else:
                worksheet_FINAL.column_dimensions[column].width = 12

        for column in ascii_uppercase:
            if column == 'A':
                worksheet_CHECK.column_dimensions[column].width = 36
                # self.worksheet.cell.font = Font(name='Arial',size=10)
            elif column == 'B':
                worksheet_CHECK.column_dimensions[column].width = 30
            else:
                worksheet_CHECK.column_dimensions[column].width = 12

        # self.workbook.save(filename=self.excel_path)  # 按照excel_path，保存路径。这里的filename=可要可不要
        self.workbook.save(filename=r"C:\Users\lxi\OneDrive - Sigma Technology\Desktop\eric-pc-gateway-1.47.0-44\eric-pc-gateway-1.47.0-44\Output_Final.xlsx")


    def replace(self,low_priority,high_priority):
        if high_priority.cell(2,1).value is None:
            pass
        else:
            final_data_list=[]
            to_be_checked=[]
            count_i=0
            flag=0
            for i in range(1, high_priority.max_row+1):
                each_podname_high = high_priority.cell(row=i, column=1).value
                count_i = count_i +1  # 避免k因为each_container_low一样而反复循环
                # print(count_i)
                count_j = 0
                done_flag=0
                for j in range(1, low_priority.max_row+1):
                    if same_service_container_flag ==1:
                        pass
                    else:
                        count_j = count_j+1
                        # print(count_j)
                        count_k=0
                        each_podname_low = low_priority.cell(row=j, column=1).value
                        # print(each_podname_low)
                        if each_podname_high == 'Chart Name':
                            continue
                        elif each_podname_high == each_podname_low:
                            each_container_high = high_priority.cell(row=i, column=2).value
                            # print(each_container_high)
                            if flag == 1:
                                continue
                            else:

                                if each_container_high is None: #high contain为空，得check
                                    # low_priority.cell(row=i,column=10).value = "Check the empty container" #不能，这在high里没有这行，没有地方可以写，得贴到Check List里
                                    to_be_checked.append(each_podname_high)
                                    to_be_checked.append(each_container_high)
                                    to_be_checked.append(high_priority.cell(row=i, column=3).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=4).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=5).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=6).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=7).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=8).value)
                                    to_be_checked.append(high_priority.cell(row=i, column=9).value)
                                else:
                                    same_service_container_flag=0
                                    for k in range(1, low_priority.max_row+1):
                                        count_k = count_k+1
                                        # print(count_k)
                                        each_container_low = low_priority.cell(row=k, column=2).value
                                        # print(each_container_low)
                                        if each_container_high == 'Container Name':
                                            continue
                                        # elif each_container_high != each_container_low:
                                        #     to_be_checked.append(each_podname_high)
                                        #     to_be_checked.append(each_container_high)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=3).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=4).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=5).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=6).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=7).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=8).value)
                                        #     to_be_checked.append(high_priority.cell(row=i, column=9).value)
                                        elif each_container_high == each_container_low and done_flag == 0:
                                            same_service_container_flag = 1
                                            done_flag=1
                                            if high_priority.cell(row=i,column=3).value is not None:
                                                low_priority.cell(row=j,column=3).value = high_priority.cell(row=i,column=3).value
                                                low_priority.cell(row=j, column=3).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=4).value is not None:
                                                low_priority.cell(row=j,column=4).value = high_priority.cell(row=i,column=4).value
                                                low_priority.cell(row=j, column=4).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=5).value is not None:
                                                low_priority.cell(row=j,column=5).value = high_priority.cell(row=i,column=5).value
                                                low_priority.cell(row=j, column=5).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=6).value is not None:
                                                low_priority.cell(row=j,column=6).value = high_priority.cell(row=i,column=6).value
                                                low_priority.cell(row=j, column=6).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=7).value is not None:
                                                low_priority.cell(row=j,column=7).value = high_priority.cell(row=i,column=7).value
                                                low_priority.cell(row=j, column=7).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=8).value is not None:
                                                low_priority.cell(row=j,column=8).value = high_priority.cell(row=i,column=8).value
                                                low_priority.cell(row=j, column=8).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                            else:
                                                pass
                                            if high_priority.cell(row=i,column=9).value is not None:
                                                low_priority.cell(row=j,column=9).value = high_priority.cell(row=i,column=9).value
                                                low_priority.cell(row=j, column=9).fill=PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色

                                            else:
                                                pass

                                        else:
                                            pass

                        else:
                            pass

            low_priority.append(to_be_checked)

                                # data_list.append(each_podname_high)
                                # data_list.append(each_container_high)
                                # if high_priority.cell(row=i, column=3).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=3).value)
                                #     low_priority.cell(row=i, column=3).fill = PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=3).value)
                                #
                                # if high_priority.cell(row=i, column=4).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=4).value)
                                #     low_priority.cell(row=i, column=4).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=4).value)
                                #
                                # if high_priority.cell(row=i, column=5).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=5).value)
                                #     low_priority.cell(row=i, column=5).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=5).value)
                                #
                                # if high_priority.cell(row=i, column=6).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=6).value)
                                #     low_priority.cell(row=i, column=6).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=6).value)
                                #
                                # if high_priority.cell(row=i, column=7).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=7).value)
                                #     low_priority.cell(row=i, column=7).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=7).value)
                                #
                                # if high_priority.cell(row=i, column=8).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=8).value)
                                #     low_priority.cell(row=i, column=8).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=8).value)
                                #
                                # if high_priority.cell(row=i, column=9).value is not None:
                                #     data_list.append(high_priority.cell(row=i, column=9).value)
                                #     low_priority.cell(row=i, column=9).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
                                # else:
                                #     data_list.append(low_priority.cell(row=j, column=9).value)






        #
        # #i是high pod，j是low pod，k是low的container
        # for i in range(1, high_priority.max_row+1):
        #     each_podname_high = high_priority.cell(row=i, column=1).value
        #     count = 0 #避免k因为each_container_low一样而反复循环
        #     for j in range(1, low_priority.max_row+1):
        #         each_podname_low = low_priority.cell(row=j, column=1).value
        #         if each_podname_high == 'Chart Name':
        #             continue
        #         elif each_podname_high == each_podname_low:
        #             each_container_high = high_priority.cell(row=i, column=2).value
        #             # print(each_container_high)
        #             for k in range(1, low_priority.max_row+1):
        #                 data_list =[]
        #                 each_container_low = low_priority.cell(row=k, column=2).value
        #                 # print(each_container_low)
        #                 if each_container_high == 'Container Name' or each_container_low == 'Container Name':
        #                     continue
        #                 elif each_container_high == each_container_low and count == 0:
        #                     count = count + 1
        #                     data_list.append(each_podname_high)
        #                     data_list.append(each_container_high)
        #                     if high_priority.cell(row=i, column=3).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=3).value)
        #                         # worksheet_FINAL.cell(row=i, column=3).fill = PatternFill(fill_type='solid', fgColor="FFBB02") #如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=3).value)
        #
        #                     if high_priority.cell(row=i, column=4).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=4).value)
        #                         # worksheet_FINAL.cell(row=i, column=4).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=4).value)
        #
        #                     if high_priority.cell(row=i, column=5).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=5).value)
        #                         # worksheet_FINAL.cell(row=i, column=5).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=5).value)
        #
        #                     if high_priority.cell(row=i, column=6).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=6).value)
        #                         # worksheet_FINAL.cell(row=i, column=6).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=6).value)
        #
        #                     if high_priority.cell(row=i, column=7).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=7).value)
        #                         # worksheet_FINAL.cell(row=i, column=7).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=7).value)
        #
        #                     if high_priority.cell(row=i, column=8).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=8).value)
        #                         # worksheet_FINAL.cell(row=i, column=8).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=8).value)
        #
        #                     if high_priority.cell(row=i, column=9).value is not None:
        #                         data_list.append(high_priority.cell(row=i, column=9).value)
        #                         # worksheet_FINAL.cell(row=i, column=9).fill = PatternFill(fill_type='solid',fgColor="FFBB02")  # 如果覆盖了就变色
        #                     else:
        #                         data_list.append(low_priority.cell(row=j, column=9).value)
        #                     # print('\n')
        #
        #
        #                     # data_list.append(each_podname_high)
        #                     # data_list.append(each_container_high)
        #                     # data_list.append(high_priority.cell(row=i, column=3).value)
        #                     # color = PatternFill(fill_type='solid',fgColor="FFBB02")
        #                     # high_priority.cell(row=i, column=3).fill = color
        #                     # data_list.append(high_priority.cell(row=i, column=4).value)
        #                     # data_list.append(high_priority.cell(row=i, column=5).value)
        #                     # data_list.append(high_priority.cell(row=i, column=6).value)
        #                     # data_list.append(high_priority.cell(row=i, column=7).value)
        #                     # data_list.append(high_priority.cell(row=i, column=8).value)
        #                     # data_list.append(high_priority.cell(row=i, column=9).value)
        #
        #                     # # break #怎样跳出k的循环？
        #                     print(data_list)
        #                     self.worksheet.append(data_list)
        #                     # print('\n')
        #                 else:
        #                     pass
        #             # self.worksheet.append(data_list)
        #             # final_data_list.append((data_list))
        #             # print(final_data_list)
        #         else:
        #             pass
        # # final_data_list.append(data_list)
        # # print(final_data_list)


        # self.worksheet.append(final_data_list)


















#如果从其他的Python文件里读取值：
    # def replace(self,low_priority,high_priority):
    #     final_data_list=[]
    #     for podname in high_priority.data_list[0]: #遍历User里的第一列
    #         if low_priority.data_list[0] = high_priority.data_list[0] and low_priority.data_list[1] = high_priority.data_list[1]: #如果User和PCG的两列都相等（podname和container name），用User的
    #             final_data_list.append(high_priority.data_list[podname]) #看看append(data_list[podname])出来是啥
    #         else:
    #             final_data_list.append(low_priority.data_list[podname])
    #     self.worksheet.append(final_data_list)


if __name__ == '__main__':
    extracted = FINAL_Extracted.Final_Extract_Data(FINAL_settings.PATH,FINAL_settings.excel_path)
    extracted.action(FINAL_settings.excel_path)